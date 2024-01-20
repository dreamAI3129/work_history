from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support import ui
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import *
import json
import func
import random
import string
import os
from datetime import datetime
import sys
from time import sleep
import pandas as pd
import openpyxl

def submit():
    with open('verified_emails.json', 'r') as file:
        accounts = json.load(file)
    for account in accounts:
        if account['applied'] == False:
            name = account['profile']
            email = account['email']
            password = "qwe123!@#"

            with open(f'profiles/{name}.json', 'r') as fp:
                profile = json.load(fp)

            driver = webdriver.Chrome()
            driver.maximize_window()
            action = webdriver.ActionChains(driver)

            func.login(driver, email, password)
            func.get_started(driver)
            
            account['applied'] = True
            with open('verified_emails.json', 'w') as file:
                json.dump(accounts, file)

            func.select_experience(driver)
            func.select_what_is_my_goal(driver)
            func.select_work_preference(driver)
            func.select_manualmode(driver)
            func.add_professional(driver, profile['professional'])
            func.add_experience(driver, profile['work_experience'])
            func.add_education(driver, profile['education'], action)
            func.add_language(driver, profile['languages'])
            func.add_skills(driver, profile['skills'])
            func.add_overview(driver, profile['overview'])
            func.add_service(driver, profile['services'])
            func.add_rate(driver, profile['hour_rate'])
            func.add_photo_others(driver, profile['photo_others'], action)
            func.submit_profile(driver)
            func.notification_imme(driver)
            #func.agency_name(driver, profile['agency'])
            driver.quit()

if __name__ == "__main__":
    driver_email = webdriver.Chrome()
    driver_email.get('https://yopmail.com/')
    while True :
        response_1 = input('Would you like to create new emails? (y/n) :')
        response_2 = input('Would you like to apply your profile to new emails? (y/n) :')
        response_3 = input('input email excel file name :')
        response_3=response_3+".xlsx"
        # response_1 = 'y'
        # response_2 = 'y'
        if response_1 == 'y':
            count = int(input('Enter number of accounts for each profile. '))
            if count > 0:
                subfix = '.'
                while True:
                    text = input(f'Enter profiles to be used{subfix} ')
                    profiles = text.split(' ')
                    exists = True
                    for name in profiles:
                        if not os.path.exists(f'profiles/{name}.json'):
                            exists = False
                    if exists:
                        break
                    else:
                        subfix = ' again.'
                    sleep(1)
                profiles = text.split(' ')
                excel = openpyxl.load_workbook(response_3)
                sheet = excel.active
                excel_file = response_3
                # rowCount = sheet.max_row + 1
                rowCount = count + 1
                emails = []
                delta =[]
                stack_tech=text
                row=1
                # driver_email = webdriver.Chrome()
                # driver_email.get('https://yopmail.com/')
                while row < rowCount:
                
                    email = sheet.cell(row=row, column=1).value
                    print(sheet.cell(row=row, column=2).value)
                    if sheet.cell(row=row, column=2).value!="verified" :
                        if sheet.cell(row=row,column=2).value=="not available":
                            rowCount=rowCount+1
                        else:
                            emails.append({"name" : email, "verified" : False})
                            delta.append(row)
                            temp_email=[]
                            temp_email.append({"name" : email, "verified" : False})

                            # if func.verify_email(driver_email,temp_email, profiles, 1) == "success" :

                            # sheet.cell(row=row, column=2).value="verified"
                            # sheet.cell(row=row, column=4).value=text
                            # excel.save(excel_file)
                    
                    elif sheet.cell(row=row, column=2).value=="verified":
                        rowCount=rowCount+1
                        
                    row=row+1
                print(emails)
                # driver_email.quit()
                # emails = func.get_email(count)
                # func.verify_email(driver_email,emails, profiles, count,sheet,excel,stack_tech,delta,excel_file)

                # def verify_email(driver_email,emails : list, profiles : list, count : int,sheet,excel,stack_tech,delta:int,excel_file):
                # driver_email = webdriver.Chrome()
                # driver_email.get('https://yopmail.com/')

                accounts = []
                for count in range(count):
                    print("count------")
                    print(count)
                    for profile_index, profile_name in enumerate(profiles):
                        with open(f'profiles/{profile_name}.json', 'r') as file:
                            profile = json.load(file)
                        country = profile['photo_others']['country']
                        first_names = profile['first_name']
                        last_names = profile['last_name']
                        yet = True
                        while yet:
                            NotVerifedEmailExist = False
                            for index, item in enumerate(emails):
                                if item['verified'] == False:
                                    email = item['name']
                                    selected_email_index = index
                                    NotVerifedEmailExist = True
                                    break
                            if NotVerifedEmailExist:
                                driver = webdriver.Chrome()
                                driver.get('https://www.upwork.com/nx/signup/?dest=home')

                                func.wait_url(driver, 'https://www.upwork.com/nx/signup/?dest=home')
                                driver.find_element(By.ID, "button-box-4").click()     
                                sleep(0.3)
                                driver.find_element(By.CSS_SELECTOR, "button[data-qa='btn-apply']").click()    
                                sleep(1)

                                first_name = first_names[random.randint(0, len(first_names)-1)]
                                last_name = last_names[random.randint(0, len(last_names)-1)]
                                driver.find_element(By.ID, "first-name-input").send_keys(first_name)
                                sleep(0.3)
                                driver.find_element(By.ID, "last-name-input").send_keys(last_name)
                                sleep(0.3)
                                driver.find_element(By.ID, 'redesigned-input-email').send_keys(email)
                                sleep(0.3)
                                all_chars = string.ascii_letters + string.digits
                                # password = random.choice(string.ascii_letters) + random.choice(string.digits)
                                # for _ in range(10):
                                #     password += random.choice(all_chars)
                                # password = ''.join(random.sample(password, len(password)))
                                password="qwe123!@#"
                                driver.find_element(By.ID, "password-input").send_keys(password)
                                sleep(0.3)
                                span = driver.find_element(By.CLASS_NAME, 'up-dropdown-toggle-title').find_element(By.TAG_NAME, 'span')
                                driver.execute_script("arguments[0].innerText = arguments[1];", span, country)
                                sleep(0.3)
                                driver.execute_script('document.querySelectorAll("span.up-checkbox-replacement-helper")[1].click()')
                                sleep(0.3)
                                func.find_element(driver, By.ID, 'button-submit-form').click()
                                sleep(3)

                                try:
                                    driver.find_element(By.CLASS_NAME, 'up-alert')
                                    print('rejected', email)
                                    driver.find_element(By.CSS_SELECTOR, "a[href='/ab/account-security/login']")
                                    print('email is rejected', email)
                                    emails[selected_email_index]['verified'] = True
                                    sheet.cell(row=delta[selected_email_index], column=2).value="not available"
                                    excel.save(excel_file)
                                except:
                                    
                                    asdf=driver.find_elements(By.CSS_SELECTOR, "a[href='/ab/account-security/login']")
                                    if len(asdf)==4:
                                        print('email is already used, rejected', email)
                                        emails[selected_email_index]['verified'] = True
                                        sheet.cell(row=delta[selected_email_index], column=5).value="Already signed, but not sure that is verified"
                                        excel.save(excel_file)
                                    else:
                                        if sheet.cell(row=delta[selected_email_index], column=2).value!="verified" :
                                            sheet.cell(row=delta[selected_email_index], column=2).value="verified"
                                            sheet.cell(row=delta[selected_email_index], column=4).value=stack_tech
                                            excel.save(excel_file)
                                            print("row number---")
                                            print(delta[selected_email_index])

                                            func.wait_url(driver, 'https://www.upwork.com/nx/signup/please-verify')
                                            print("before pressing refresh button on yopmail")
                                            driver_email.refresh()
                                            print("yopmail refreshed")
                                            sleep(2)
                                            if count == 0 and profile_index == 0:                        
                                                func.find_element(driver_email, By.CLASS_NAME, 'ycptinput').send_keys(email)
                                                while True:
                                                    response = input('passed CAPTCHA? (y/n)')
                                                    if response =='y':
                                                        break
                                                    sleep(1)
                                            while True:
                                                print('waiting...')
                                                try:
                                                    driver_email.switch_to.frame('ifmail')
                                                    a_tags = func.find_elements(driver_email, By.TAG_NAME, 'a')
                                                    found_verify_url = False
                                                    for tag in a_tags:
                                                        if 'signup/verify-email' in tag.get_attribute('href'):
                                                            driver.get(tag.get_attribute('href'))
                                                            found_verify_url = True
                                                    if found_verify_url:
                                                        break                                    
                                                except:
                                                    pass
                                                sleep(1)
                                            func.wait_url(driver, 'https://www.upwork.com/nx/create-profile/')
                                            accounts.append({"profile" : profile_name, "email" : email, "password" : password, "name" : first_name + ' ' + last_name, "applied"  : False})
                                            with open('verified_emails.json', 'w') as file:
                                                json.dump(accounts, file)
                                            yet = False      
                                            print('created', email)
                                            emails[selected_email_index]['verified'] = True

                                    
                                    

                                driver.quit()
                            else:
                                yet = False
                                print('emails was verified')
                            sleep(1)
                        if not NotVerifedEmailExist:
                            break
                # driver_email.quit()        
                print('emails was verified.')
                # return "success"




        if response_2 == 'y':
            submit()
        print('done')
